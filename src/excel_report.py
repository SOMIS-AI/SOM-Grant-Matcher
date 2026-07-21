"""
Excel Match Report
==================
Builds an .xlsx workbook for the daily/weekly grant digests:
  • a "Summary" tab listing every matched grant, and
  • one tab per grant with the grant's metadata + a table of matched faculty
    (Faculty, Department, Email, Confidence, Match Type, Matched Keywords, Profile).

Returned as bytes for attaching to the SendGrid email. Degrades gracefully:
build_workbook_bytes() returns None if openpyxl isn't installed.
"""

import logging
import re
from datetime import datetime
from io import BytesIO

logger = logging.getLogger(__name__)


# ── Match accessors (handle both Match namedtuples and plain dicts) ───────────

def _attr(m, name, default=""):
    v = getattr(m, name, None)
    if v is None and isinstance(m, dict):
        v = m.get(name)
    return default if v is None else v


def _conf(m) -> int:
    c = _attr(m, "confidence_score", None)
    if c is None:
        c = _attr(m, "match_score", 0)
    try:
        return int(c)
    except (TypeError, ValueError):
        return 0


def _keywords(m) -> str:
    kws = _attr(m, "matched_keywords", []) or []
    return ", ".join(kws)


# ── Sheet-name sanitisation (Excel: <=31 chars, no []:*?/\, must be unique,
#    must not START or END with an apostrophe) ─────────────────────────────────

def _safe_sheet_name(title: str, used: set) -> str:
    name = re.sub(r"[\[\]:*?/\\]", " ", title or "Grant").strip() or "Grant"
    name = name[:31].strip("'").strip() or "Grant"   # edge apostrophes are illegal
    base = name
    n = 2
    while name.lower() in used:
        suffix = f" ({n})"
        name = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(name.lower())
    return name


def _sheet_ref(sheet_name: str) -> str:
    """Internal hyperlink target for a sheet. Apostrophes INSIDE a quoted sheet
    reference must be doubled ('' ) or Excel reports \"Reference isn't valid\" —
    which hit every Summary link for titles like \"Alzheimer's Disease ...\"."""
    return "#'{}'!A1".format(sheet_name.replace("'", "''"))


_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _text_cell(ws, row, column, value):
    """Write a TEXT cell, never a formula. openpyxl treats any string starting
    with '=' as a live formula, so a scraped grant title crafted (or accidentally
    formed) as '=HYPERLINK(...)' would execute in a report emailed to leadership.
    Forces string storage for the risky prefixes (=, +, -, @)."""
    cell = ws.cell(row=row, column=column, value=value)
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        cell.data_type = "s"
    return cell


def build_workbook_bytes(matched_results: list, run_date: str, label: str = "Daily"):
    """
    Build the .xlsx and return its bytes, or None if openpyxl is unavailable
    or there are no results.
    """
    if not matched_results:
        return None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel attachment.")
        return None

    NAVY = "1A2E45"
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color="FFFFFF")
    title_font  = Font(bold=True, size=14, color=NAVY)
    label_font  = Font(bold=True, color="555555")
    link_font   = Font(color="1A4B6E", underline="single")

    # Sort grants by best (first) match confidence, descending.
    def _best(r):
        return _conf(r["matches"][0]) if r.get("matches") else 0
    results = sorted(matched_results, key=lambda r: (_best(r), len(r.get("matches", []))), reverse=True)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    # "history" reserved: Excel treats a sheet literally named History as the
    # shared-workbook history sheet and prompts a repair on open.
    used_names = {"summary", "history"}

    total_grants = len(results)
    total_matches = sum(len(r.get("matches", [])) for r in results)

    summary["A1"] = f"UMSOM Grant Matcher — {label} Match Report"
    summary["A1"].font = title_font
    summary["A2"] = f"Run date: {run_date}"
    summary["A3"] = f"{total_grants} grants matched to {total_matches} faculty members"

    hdr = ["#", "Grant Title", "Agency", "Matches", "Avg Confidence", "Deadline"]
    for c, h in enumerate(hdr, 1):
        cell = summary.cell(row=5, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Pre-compute sheet names so the Summary can hyperlink to each grant tab.
    sheet_names = [_safe_sheet_name(r["grant"].get("title", "Grant"), used_names) for r in results]

    for idx, (r, sheet_name) in enumerate(zip(results, sheet_names), 1):
        grant = r["grant"]
        matches = r.get("matches", [])
        confs = [_conf(m) for m in matches]
        avg_conf = round(sum(confs) / len(confs)) if confs else 0

        row = 5 + idx
        summary.cell(row=row, column=1, value=idx)
        title_cell = _text_cell(summary, row, 2, grant.get("title", ""))
        # Internal hyperlink to the grant's tab (apostrophe-safe reference)
        title_cell.hyperlink = _sheet_ref(sheet_name)
        title_cell.font = link_font
        _text_cell(summary, row, 3, grant.get("agency", ""))
        summary.cell(row=row, column=4, value=len(matches))
        summary.cell(row=row, column=5, value=f"{avg_conf}%")
        _text_cell(summary, row, 6, grant.get("close_date", "") or "N/A")

    for col, width in zip("ABCDEF", (5, 60, 32, 10, 14, 14)):
        summary.column_dimensions[col].width = width

    # ── One sheet per grant ──────────────────────────────────────────────────
    for r, sheet_name in zip(results, sheet_names):
        grant = r["grant"]
        matches = sorted(r.get("matches", []), key=lambda m: -_conf(m))
        ws = wb.create_sheet(title=sheet_name)

        _text_cell(ws, 1, 1, grant.get("title", ""))
        ws["A1"].font = title_font

        meta = [
            ("Agency:",        grant.get("agency", "") or "N/A"),
            ("Grant Number:",  grant.get("number", "") or "N/A"),
            ("Award Ceiling:", grant.get("award_ceiling", "") or "N/A"),
            ("Open Date:",     grant.get("open_date", "") or "N/A"),
            ("Close Date:",    grant.get("close_date", "") or "N/A"),
        ]
        rownum = 2
        for lbl, val in meta:
            ws.cell(row=rownum, column=1, value=lbl).font = label_font
            _text_cell(ws, rownum, 2, str(val))
            rownum += 1
        # Grant page link
        ws.cell(row=rownum, column=1, value="Grant Page:").font = label_font
        link = grant.get("link", "")
        gp = ws.cell(row=rownum, column=2, value=link)
        if link:
            gp.hyperlink = link
            gp.font = link_font
        rownum += 2

        back = ws.cell(row=rownum, column=1, value="← Back to Summary")
        back.hyperlink = "#'Summary'!A1"
        back.font = link_font
        rownum += 2

        # Faculty table
        cols = ["Faculty", "Department", "Email", "Confidence", "Match Type",
                "Matched Keywords", "Profile"]
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=rownum, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        rownum += 1

        for m in matches:
            _text_cell(ws, rownum, 1, _attr(m, "faculty_name", ""))
            _text_cell(ws, rownum, 2, _attr(m, "faculty_department", ""))
            _text_cell(ws, rownum, 3, _attr(m, "faculty_email", ""))
            ws.cell(row=rownum, column=4, value=f"{_conf(m)}%")
            ws.cell(row=rownum, column=5, value=_attr(m, "match_type", "keyword"))
            _text_cell(ws, rownum, 6, _keywords(m))
            furl = _attr(m, "faculty_url", "")
            pcell = ws.cell(row=rownum, column=7, value="View Profile" if furl else "")
            if furl:
                pcell.hyperlink = furl
                pcell.font = link_font
            rownum += 1

        for col, width in zip("ABCDEFG", (26, 30, 30, 12, 14, 50, 14)):
            ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def report_filename(run_date_dt: datetime = None, label: str = "Daily") -> str:
    d = (run_date_dt or datetime.utcnow()).strftime("%Y-%m-%d")
    # Include the Daily/Weekly label so the two emails sent on the weekly weekday
    # produce distinct filenames (otherwise saving both overwrites the first).
    lbl = "".join(c for c in (label or "Daily") if c.isalnum()) or "Daily"
    return f"UMSOM_Grant_Matches_{lbl}_{d}.xlsx"
